#!/usr/bin/env python3
"""
Generate SVG thermal label previews for SATO productos intermedios.
Integrates with HTML diagrams for clickable label overlays.
"""

import openpyxl
import os
import re
from datetime import datetime
from pathlib import Path

# Paths
EXCEL_PATH = "/sessions/confident-cool-maxwell/mnt/IMPRESORA PRODUCTOS INTERMEDIOS/RIA_Referencia_Etiquetas_SATO.xlsx"
ETIQUETAS_DIR = "/sessions/confident-cool-maxwell/mnt/calidad/etiquetas"
HTML_DIR = "/sessions/confident-cool-maxwell/mnt/calidad"

# Load Excel data
def load_label_data():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    caldos = []
    salsas = []

    if "Caldos" in wb.sheetnames:
        ws = wb["Caldos"]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if row[0]:  # producto name
                caldos.append({
                    'producto': row[0],
                    'ingredientes': row[1] if len(row) > 1 else '',
                    'prefijo_lote': row[2] if len(row) > 2 else '',
                    'vida_util': row[3] if len(row) > 3 else 21,
                    'conservacion': row[4] if len(row) > 4 else '',
                })

    if "Salsas" in wb.sheetnames:
        ws = wb["Salsas"]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if row[0]:  # producto name
                salsas.append({
                    'producto': row[0],
                    'ingredientes': row[1] if len(row) > 1 else '',
                    'prefijo_lote': row[2] if len(row) > 2 else '',
                    'vida_util': row[3] if len(row) > 3 else 21,
                    'conservacion': row[4] if len(row) > 4 else '',
                })

    return caldos, salsas

def format_ingredients_with_bold(text):
    """
    Extract allergens (text in ALL CAPS) and return a list of
    (text, is_allergen) tuples for proper SVG rendering.
    """
    if not text:
        return []

    # Split by common delimiters and rebuild with allergen detection
    parts = []
    current = ""
    i = 0
    while i < len(text):
        # Check for ALL CAPS words (allergens)
        match = re.match(r'[A-Z]{2,}(?:\s+[A-Z]{2,})*', text[i:])
        if match:
            allergen = match.group(0)
            if current:
                parts.append((current, False))
                current = ""
            parts.append((allergen, True))
            i += len(allergen)
        else:
            current += text[i]
            i += 1

    if current:
        parts.append((current, False))

    return parts

def generate_svg_label(producto, ingredientes, prefijo_lote, conservacion):
    """
    Generate SVG thermal label (40mm x 80mm = 400 x 800 units at 10dpi).
    Realistic thermal label design with proper formatting.
    """

    # SVG dimensions (thermal label: 40mm x 80mm at 10 dots/mm = 400x800)
    width = 400
    height = 800

    svg_lines = [
        f'<?xml version="1.0" encoding="UTF-8"?>',
        f'<svg width="{width}" height="{height}" viewBox="0 0 {width} {height}" xmlns="http://www.w3.org/2000/svg">',
        # White background
        f'<rect width="{width}" height="{height}" fill="white" stroke="#333" stroke-width="2" rx="8"/>',
    ]

    y_pos = 30

    # Product name - large and bold at top
    svg_lines.extend([
        f'<text x="200" y="{y_pos}" font-size="28" font-weight="bold" text-anchor="middle" fill="black">',
        f'  {producto}',
        f'</text>',
    ])

    y_pos += 50

    # Ingredients section
    svg_lines.append(f'<text x="15" y="{y_pos}" font-size="8" font-weight="bold" fill="#333">INGREDIENTES:</text>')
    y_pos += 12

    # Parse and render ingredients with allergen bolding
    if ingredientes:
        # Simple text wrapping at ~60 chars per line for small font
        lines = []
        words = ingredientes.split()
        current_line = []
        current_length = 0

        for word in words:
            if current_length + len(word) > 50:
                if current_line:
                    lines.append(' '.join(current_line))
                current_line = [word]
                current_length = len(word)
            else:
                current_line.append(word)
                current_length += len(word) + 1

        if current_line:
            lines.append(' '.join(current_line))

        # Render each line with allergen detection
        for line in lines[:6]:  # Max 6 lines for ingredients
            # Simple approach: render line, highlight CAPS as bold
            svg_lines.append(f'<text x="15" y="{y_pos}" font-size="7" fill="black" font-family="monospace">')

            # Process line for allergens
            parts = []
            current = ""
            for char in line:
                if char == ' ' or not char.isupper():
                    if current and all(c.isupper() or c in ' -(),' for c in current):
                        # Might be allergen
                        parts.append(('ALLERGEN', current))
                        current = ""
                    if char == ' ':
                        if current:
                            parts.append(('NORMAL', current + ' '))
                            current = ""
                        else:
                            if parts and parts[-1][0] == 'NORMAL':
                                parts[-1] = ('NORMAL', parts[-1][1] + ' ')
                            else:
                                parts.append(('NORMAL', ' '))
                    else:
                        current += char
                else:
                    current += char

            if current:
                if all(c.isupper() or c in ' -(),' for c in current):
                    parts.append(('ALLERGEN', current))
                else:
                    parts.append(('NORMAL', current))

            # Render with tspans for bold allergens
            svg_lines[-1] = svg_lines[-1][:-7]  # Remove closing tag

            for ptype, ptext in parts:
                if ptype == 'ALLERGEN':
                    svg_lines.append(f'<tspan font-weight="bold" fill="#c00">{ptext}</tspan>')
                else:
                    svg_lines.append(f'<tspan fill="black">{ptext}</tspan>')

            svg_lines.append('</text>')
            y_pos += 10

    y_pos += 10

    # Conservation
    if conservacion:
        svg_lines.extend([
            f'<text x="15" y="{y_pos}" font-size="8" font-weight="bold" fill="#333">CONSERVACIÓN:</text>',
        ])
        y_pos += 10
        svg_lines.append(f'<text x="15" y="{y_pos}" font-size="7" fill="black">{conservacion}</text>')
        y_pos += 15

    # Batch/Lote line
    today = datetime.now()
    batch_code = f"{prefijo_lote}{today.strftime('%d%m%y')}"
    svg_lines.extend([
        f'<text x="15" y="{y_pos}" font-size="9" font-weight="bold" fill="black">',
        f'  LOTE: {batch_code}',
        f'</text>',
    ])
    y_pos += 16

    # Elaboration date
    elab_date = today.strftime('%d/%m/%Y')
    svg_lines.append(f'<text x="15" y="{y_pos}" font-size="8" fill="black">ELAB: {elab_date}</text>')
    y_pos += 12

    # Expiration date (vida_util days from now)
    from datetime import timedelta
    cad_date = (today + timedelta(days=21)).strftime('%d/%m/%Y')
    svg_lines.append(f'<text x="15" y="{y_pos}" font-size="8" fill="black">CAD: {cad_date}</text>')

    # Close SVG
    svg_lines.append('</svg>')

    return '\n'.join(svg_lines)

def sanitize_filename(name):
    """Convert product name to safe filename."""
    name = name.lower()
    name = re.sub(r'[^\w\s-]', '', name)
    name = re.sub(r'[\s-]+', '_', name)
    return name

def main():
    caldos, salsas = load_label_data()

    print(f"Loaded {len(caldos)} caldos and {len(salsas)} salsas")

    # Mapping of producto names to diagrams
    diagram_mapping = {
        'Caldo de Pescado': 'dark_pescado.html',
        'Caldo Negro': 'dark_negro.html',
        'Caldo de Pollo': 'dark_pollo.html',
        'Caldo de Verduras': 'dark_verduras.html',
        'Caldo de Risotto': 'dark_risotto.html',
        'Salsa de Pescado': 'dark_salsa_pescado.html',
        'Salsa Negra': 'dark_salsa_negra.html',
        'Salsa de Verduras': 'dark_salsa_verduras.html',
        'Salsa de Pollo': 'dark_salsa_pollo.html',
        'Salsa Tartufata': 'dark_salsa_tartufata.html',
        'Salsa de Boletus': 'dark_salsa_boletus.html',
        'Salsa de Setas': 'dark_salsa_setas.html',
        'Salsa de Queso': 'dark_salsa_queso.html',
        'Salsa Iberico': 'dark_salsa_iberica.html',
    }

    generated = {}

    # Generate SVG labels for all products
    for caldo in caldos:
        nombre = caldo['producto']
        filename = sanitize_filename(nombre) + '.svg'
        filepath = os.path.join(ETIQUETAS_DIR, filename)

        svg_content = generate_svg_label(
            nombre,
            caldo['ingredientes'],
            caldo['prefijo_lote'],
            caldo['conservacion']
        )

        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(svg_content)

        generated[nombre] = filename
        print(f"✓ Generated: {filename}")

    for salsa in salsas:
        nombre = salsa['producto']
        filename = sanitize_filename(nombre) + '.svg'
        filepath = os.path.join(ETIQUETAS_DIR, filename)

        svg_content = generate_svg_label(
            nombre,
            salsa['ingredientes'],
            salsa['prefijo_lote'],
            salsa['conservacion']
        )

        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(svg_content)

        generated[nombre] = filename
        print(f"✓ Generated: {filename}")

    # Now integrate labels into HTML diagrams
    integrate_labels_to_html(diagram_mapping, generated)

def integrate_labels_to_html(diagram_mapping, generated):
    """Add label layer to HTML diagrams."""

    # Reverse mapping: diagram → products to label
    html_products = {}
    for prod_name, html_file in diagram_mapping.items():
        if html_file not in html_products:
            html_products[html_file] = []
        html_products[html_file].append(prod_name)

    for html_file, products in html_products.items():
        html_path = os.path.join(HTML_DIR, html_file)
        if not os.path.exists(html_path):
            print(f"⚠ HTML file not found: {html_file}")
            continue

        with open(html_path, 'r', encoding='utf-8') as f:
            content = f.read()

        # Find the script section (before closing body)
        script_insert_idx = content.rfind('</script>')
        if script_insert_idx == -1:
            print(f"⚠ No script closing found in {html_file}")
            continue

        # Locate insertion point right before first </script> closing
        insert_idx = content.rfind('<script', 0, script_insert_idx)
        if insert_idx != -1:
            # Find the closing tag
            script_close_idx = content.find('</script>', insert_idx)
            if script_close_idx != -1:
                insert_idx = script_close_idx

        # Add modal HTML and update nodes with data-etiqueta attributes
        modal_html = '''
<!-- LABEL MODAL OVERLAY -->
<div id="labelModal" style="display:none;position:fixed;top:0;left:0;right:0;bottom:0;background:rgba(0,0,0,0.8);z-index:1000;align-items:center;justify-content:center">
  <div style="position:relative;background:white;padding:20px;border-radius:12px;max-width:600px;width:90%;max-height:80vh;overflow:auto;box-shadow:0 10px 40px rgba(0,0,0,0.5)">
    <button onclick="document.getElementById('labelModal').style.display='none'" style="position:absolute;top:10px;right:10px;background:none;border:none;font-size:24px;cursor:pointer;color:#666">×</button>
    <div id="labelContent" style="margin-top:30px"></div>
  </div>
</div>

<style>
#labelModal { display: flex !important; }
.label-button {
  font-size: 14px;
  cursor: pointer;
  color: #fbbf24;
  text-decoration: underline;
  margin-left: 8px;
}
.label-button:hover { color: #fcd34d; }
</style>

<script>
function showLabel(etiquetaPath) {
  const modal = document.getElementById('labelModal');
  const content = document.getElementById('labelContent');
  fetch(etiquetaPath)
    .then(r => r.text())
    .then(svg => {
      content.innerHTML = svg;
      modal.style.display = 'flex';
    })
    .catch(err => {
      content.innerHTML = '<p style="color:red">Error loading label: ' + err + '</p>';
      modal.style.display = 'flex';
    });
}

// Close modal on background click
document.getElementById('labelModal').addEventListener('click', function(e) {
  if (e.target === this) this.style.display = 'none';
});
</script>
'''

        # Insert modal code before closing body tag
        body_close = content.rfind('</body>')
        if body_close != -1:
            content = content[:body_close] + modal_html + content[body_close:]

        # Update nodes with label data
        for prod_name in products:
            if prod_name in generated:
                etiqueta_file = generated[prod_name]
                # Find node divs with this product name in data-n attribute
                pattern = f'data-n="{re.escape(prod_name)}"'
                if pattern in content:
                    # Insert label icon and data attribute
                    # Find the matching node div and add label button
                    lines = content.split('\n')
                    new_lines = []
                    for i, line in enumerate(lines):
                        if f'data-n="{prod_name}"' in line:
                            # Add data-etiqueta attribute
                            line = line.replace('data-s=""', f'data-etiqueta="etiquetas/{etiqueta_file}" data-s=""')
                            # Find closing </div> of this node and add label button
                            for j in range(i, min(i+10, len(lines))):
                                if '</div>' in lines[j] and 'node' in lines[i]:
                                    # Insert label button before closing div
                                    lines[j] = lines[j].replace('</div>', f' <span class="label-button" onclick="showLabel(\'etiquetas/{etiqueta_file}\')">🏷️</span></div>')
                                    break
                        new_lines.append(line)

                    content = '\n'.join(new_lines)

        with open(html_path, 'w', encoding='utf-8') as f:
            f.write(content)

        print(f"✓ Updated: {html_file}")

if __name__ == '__main__':
    main()
